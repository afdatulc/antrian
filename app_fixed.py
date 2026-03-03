from flask import Flask, render_template, request, redirect, session, jsonify
import mysql.connector
from mysql.connector import Error
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import os
import smtplib
from email.message import EmailMessage
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    openpyxl = None

app = Flask(__name__)
app.secret_key = "super_secret_key"

EXCEL_FILE = "absensi.xlsx"

# =========================
# MYSQL CONNECTION
# =========================
def get_db():
    try:
        conn = mysql.connector.connect(
            host=os.environ.get("DB_HOST", "localhost"),
            user=os.environ.get("DB_USER", "root"),
            password=os.environ.get("DB_PASS", ""),
            database=os.environ.get("DB_NAME", "antrian_db")
        )
        return conn
    except Error as e:
        print(f"Error connecting to MySQL: {e}")
        return None


# =========================
# Helper: Email & Excel
# =========================
def send_email(to_address: str, subject: str, body: str):
    user = os.environ.get("EMAIL_USER")
    password = os.environ.get("EMAIL_PASS")
    if not user or not password:
        print("EMAIL_USER / EMAIL_PASS tidak diset, email dilewati")
        return
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = user
    msg["To"] = to_address
    msg.set_content(body)
    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.starttls()
            smtp.login(user, password)
            smtp.send_message(msg)
    except Exception as e:
        print(f"Error sending email: {e}")


def append_to_excel(nama, telepon, email, no_antrian, waktu):
    if openpyxl is None:
        print("openpyxl tidak terpasang; tidak menulis excel")
        return
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nama", "Telepon", "Email", "No Antrian", "Waktu Daftar"])
    ws.append([nama, telepon, email, no_antrian, waktu])
    wb.save(EXCEL_FILE)


# =========================
# INIT DATABASE
# =========================
@app.route("/init-db")
def init_db():
    conn = get_db()
    if not conn:
        return "Gagal koneksi ke MySQL"
    
    cursor = conn.cursor()
    
    try:
        # Buat tabel users
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            role VARCHAR(20) NOT NULL,
            loket INT
        )
        """)
        
        # Buat tabel antrian
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS antrian (
            id INT AUTO_INCREMENT PRIMARY KEY,
            nama VARCHAR(100) NOT NULL,
            telepon VARCHAR(20),
            email VARCHAR(100),
            no_antrian INT,
            loket INT,
            status VARCHAR(20) DEFAULT 'Menunggu',
            waktu_daftar DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """)
        
        conn.commit()
        
        # Hapus data lama sebelum seed
        cursor.execute("DELETE FROM users")
        cursor.execute("DELETE FROM antrian")
        
        # Seed users
        admin_password = generate_password_hash("admin123")
        cursor.execute(
            "INSERT INTO users (username, password_hash, role, loket) VALUES (%s, %s, %s, %s)",
            ("admin", admin_password, "admin", None)
        )
        
        for i in range(1, 6):
            loket_password = generate_password_hash("123")
            cursor.execute(
                "INSERT INTO users (username, password_hash, role, loket) VALUES (%s, %s, %s, %s)",
                (f"loket{i}", loket_password, "interviewer", i)
            )
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return "Database MySQL siap! Seed data berhasil dimasukkan."
    
    except Error as e:
        print(f"Error: {e}")
        return f"Error: {e}"


# =========================
# PUBLIC PAGES
# =========================
@app.route("/")
def home():
    return redirect("/display")


@app.route("/display")
def display():
    return render_template("display.html")


@app.route("/data-antrian")
def data_antrian():
    conn = get_db()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    cursor = conn.cursor(dictionary=True)
    
    result = {"loket": {}, "menunggu": []}
    
    try:
        # Data per loket
        for i in range(1, 6):
            cursor.execute(
                "SELECT no_antrian FROM antrian WHERE loket=%s AND status='Dipanggil' ORDER BY id DESC LIMIT 1",
                (i,)
            )
            cur = cursor.fetchone()
            result["loket"][f"loket{i}"] = cur["no_antrian"] if cur and cur["no_antrian"] is not None else "-"
        
        # Data menunggu
        cursor.execute("SELECT no_antrian FROM antrian WHERE status='Menunggu' ORDER BY id")
        rows = cursor.fetchall()
        result["menunggu"] = [row["no_antrian"] for row in rows]
        
    except Error as e:
        print(f"Error: {e}")
    finally:
        cursor.close()
        conn.close()
    
    return jsonify(result)


@app.route("/absensi", methods=["GET"])
def absensi():
    return render_template("absensi.html")


@app.route("/submit_absensi", methods=["POST"])
def submit_absensi():
    nama = request.form["nama"]
    telepon = request.form["telepon"]
    email = request.form["email"]

    conn = get_db()
    if not conn:
        return "Gagal koneksi database"
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Hitung nomor antrian berikutnya
        cursor.execute("SELECT MAX(no_antrian) as max_no FROM antrian")
        result = cursor.fetchone()
        maxno = result["max_no"] or 0
        next_no = maxno + 1

        # Insert data
        cursor.execute(
            """
            INSERT INTO antrian (nama, telepon, email, no_antrian, loket, status, waktu_daftar)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (nama, telepon, email, next_no, None, "Menunggu", datetime.now())
        )
        conn.commit()

        # Simpan ke Excel
        append_to_excel(nama, telepon, email, next_no, datetime.now().isoformat())
        
        # Kirim email
        send_email(
            email,
            "Nomor Antrian Wawancara",
            f"Terima kasih {nama},\n\nNomor antrian Anda: {next_no}\n\nSilakan tunggu panggilan di loket yang telah ditentukan."
        )

        return render_template("terimakasih.html", nama=nama, email=email, no_antrian=next_no)
    
    except Error as e:
        print(f"Error: {e}")
        return f"Error: {e}"
    finally:
        cursor.close()
        conn.close()


# =========================
# LOGIN
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form["username"]
        pw = request.form["password"]
        
        conn = get_db()
        if not conn:
            return "Gagal koneksi database"
        
        cursor = conn.cursor(dictionary=True)
        
        try:
            cursor.execute("SELECT * FROM users WHERE username=%s", (user,))
            data = cursor.fetchone()
            
            if data and check_password_hash(data["password_hash"], pw):
                session["role"] = data["role"]
                session["loket"] = data["loket"]
                session["user"] = user
                if data["role"] == "admin":
                    return redirect("/admin")
                else:
                    return redirect("/interviewer")
            return "Login gagal"
        except Error as e:
            print(f"Error: {e}")
            return f"Error: {e}"
        finally:
            cursor.close()
            conn.close()
    
    return render_template("login.html")


# =========================
# ADMIN DASHBOARD
# =========================
@app.route("/admin")
def admin():
    if session.get("role") != "admin":
        return redirect("/login")
    
    conn = get_db()
    if not conn:
        return "Gagal koneksi database"
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        cursor.execute("SELECT * FROM antrian ORDER BY id DESC")
        antrian = cursor.fetchall()
        return render_template("admin.html", data=antrian)
    except Error as e:
        print(f"Error: {e}")
        return f"Error: {e}"
    finally:
        cursor.close()
        conn.close()


# =========================
# INTERVIEWER DASHBOARD
# =========================
@app.route("/interviewer")
def interviewer():
    if session.get("role") != "interviewer":
        return redirect("/login")
    
    loket = session["loket"]
    conn = get_db()
    if not conn:
        return "Gagal koneksi database"
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        cursor.execute(
            "SELECT * FROM antrian WHERE loket=%s AND status='Dipanggil' ORDER BY id DESC LIMIT 1",
            (loket,)
        )
        called = cursor.fetchone()
        current = called["no_antrian"] if called else "-"
        return render_template("pewawancara.html", loket=loket, current=current)
    except Error as e:
        print(f"Error: {e}")
        return f"Error: {e}"
    finally:
        cursor.close()
        conn.close()


# =========================
# PANGGIL / UPDATE
# =========================
@app.route("/panggil/<int:loket>")
def panggil(loket):
    if session.get("role") != "admin":
        return redirect("/login")
    
    conn = get_db()
    if not conn:
        return redirect("/admin")
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Cari antrian menunggu berikutnya
        cursor.execute("SELECT * FROM antrian WHERE status='Menunggu' ORDER BY id LIMIT 1")
        row = cursor.fetchone()
        
        if row:
            cursor.execute(
                "UPDATE antrian SET status='Dipanggil', loket=%s WHERE id=%s",
                (loket, row["id"])
            )
            conn.commit()
    except Error as e:
        print(f"Error: {e}")
    finally:
        cursor.close()
        conn.close()
    
    return redirect("/admin")


@app.route("/panggil-saya")
def panggil_saya():
    if session.get("role") != "interviewer":
        return redirect("/login")
    
    loket = session["loket"]
    conn = get_db()
    if not conn:
        return redirect("/interviewer")
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        cursor.execute("SELECT * FROM antrian WHERE status='Menunggu' ORDER BY id LIMIT 1")
        row = cursor.fetchone()
        
        if row:
            cursor.execute(
                "UPDATE antrian SET status='Dipanggil', loket=%s WHERE id=%s",
                (loket, row["id"])
            )
            conn.commit()
    except Error as e:
        print(f"Error: {e}")
    finally:
        cursor.close()
        conn.close()
    
    return redirect("/interviewer")


@app.route("/tambah", methods=["POST"])
def tambah():
    if session.get("role") != "admin":
        return redirect("/login")
    
    nama = request.form["nama"]
    email = request.form["email"]
    loket = request.form["loket"]
    
    conn = get_db()
    if not conn:
        return redirect("/admin")
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        cursor.execute("SELECT MAX(no_antrian) as max_no FROM antrian")
        result = cursor.fetchone()
        maxno = result["max_no"] or 0
        next_no = maxno + 1
        
        cursor.execute(
            """
            INSERT INTO antrian (nama, email, loket, status, waktu_daftar, no_antrian, telepon)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (nama, email, loket, "Menunggu", datetime.now(), next_no, "")
        )
        conn.commit()
    except Error as e:
        print(f"Error: {e}")
    finally:
        cursor.close()
        conn.close()
    
    return redirect("/admin")


@app.route("/update/<int:id>/<status>")
def update(id, status):
    conn = get_db()
    if not conn:
        return redirect(request.referrer or "/admin")
    
    cursor = conn.cursor()
    
    try:
        cursor.execute("UPDATE antrian SET status=%s WHERE id=%s", (status, id))
        conn.commit()
    except Error as e:
        print(f"Error: {e}")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(request.referrer or "/admin")


# =========================
# STATISTIK
# =========================
@app.route("/statistik")
def statistik():
    if session.get("role") != "admin":
        return redirect("/login")
    
    conn = get_db()
    if not conn:
        return "Gagal koneksi database"
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        cursor.execute("SELECT COUNT(*) as count FROM antrian WHERE status='Menunggu'")
        menunggu = cursor.fetchone()["count"]
        
        stats = {}
        for i in range(1, 6):
            cursor.execute("SELECT COUNT(*) as count FROM antrian WHERE loket=%s", (i,))
            total = cursor.fetchone()["count"]
            
            cursor.execute("SELECT COUNT(*) as count FROM antrian WHERE loket=%s AND status='Selesai'", (i,))
            selesai = cursor.fetchone()["count"]
            
            cursor.execute("SELECT COUNT(*) as count FROM antrian WHERE loket=%s AND status='Dipanggil'", (i,))
            dipanggil = cursor.fetchone()["count"]
            
            stats[i] = {"total": total, "selesai": selesai, "dipanggil": dipanggil}
        
        return render_template("statistik.html", menunggu=menunggu, statistik=stats)
    except Error as e:
        print(f"Error: {e}")
        return f"Error: {e}"
    finally:
        cursor.close()
        conn.close()


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


if __name__ == "__main__":
    app.run(debug=True)