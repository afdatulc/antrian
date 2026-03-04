from flask import Flask, render_template, request, redirect, session, jsonify, make_response
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import sqlite3
import os

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

app = Flask(__name__)
app.secret_key = "antrian_wawancara_secret_2024"

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "antrian.db")
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "absensi_wawancara.xlsx")


# =========================
# DATABASE HELPERS
# =========================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    """Initialize database tables and seed data if needed."""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            loket INTEGER,
            nama_display TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS antrian (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT NOT NULL,
            telepon TEXT,
            email TEXT,
            no_antrian INTEGER,
            loket INTEGER,
            status TEXT DEFAULT 'Menunggu',
            waktu_daftar DATETIME DEFAULT CURRENT_TIMESTAMP,
            waktu_dipanggil DATETIME,
            waktu_selesai DATETIME,
            nama_pewawancara TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS active_sessions (
            loket INTEGER PRIMARY KEY,
            session_id TEXT NOT NULL,
            login_time DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Create indexes for performance optimization
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_antrian_status ON antrian(status)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_antrian_loket ON antrian(loket)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_antrian_loket_status ON antrian(loket, status)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_antrian_no_antrian ON antrian(no_antrian)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_antrian_waktu_dipanggil ON antrian(waktu_dipanggil)")

    # Check if we need to seed
    cursor.execute("SELECT COUNT(*) FROM users")
    if cursor.fetchone()[0] == 0:
        # Seed admin
        cursor.execute(
            "INSERT INTO users (username, password_hash, role, loket, nama_display) VALUES (?, ?, ?, ?, ?)",
            ("admin", generate_password_hash("admin123"), "admin", None, "Administrator")
        )
        # Seed 25 interviewers
        for i in range(1, 26):
            cursor.execute(
                "INSERT INTO users (username, password_hash, role, loket, nama_display) VALUES (?, ?, ?, ?, ?)",
                (f"loket{i}", generate_password_hash("123"), "interviewer", i, f"Pewawancara Loket {i}")
            )

    conn.commit()
    conn.close()


# Initialize on startup
init_db()


# =========================
# HELPER FUNCTIONS
# =========================
def append_to_excel(nama, telepon, email, no_antrian, waktu):
    if openpyxl is None:
        return
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Absensi Wawancara"
        headers = ["No", "Nama Lengkap", "Telepon", "Email", "No Antrian", "Waktu Daftar", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 12

    row_num = ws.max_row + 1
    row_data = [row_num - 1, nama, telepon, email, no_antrian, waktu, "Menunggu"]
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    wb.save(EXCEL_FILE)


def get_today_str():
    return datetime.now().strftime("%Y-%m-%d")


# =========================
# PUBLIC PAGES
# =========================
@app.route("/")
def home():
    return redirect("/display")


@app.route("/display")
def display():
    return render_template("display.html")


@app.route("/absensi", methods=["GET"])
def absensi():
    return render_template("absensi.html")


@app.route("/submit_absensi", methods=["POST"])
def submit_absensi():
    nama = request.form.get("nama", "").strip()
    telepon = request.form.get("telepon", "").strip()
    email = request.form.get("email", "").strip()

    if not nama:
        return redirect("/absensi")

    conn = get_db()
    cursor = conn.cursor()

    # Get next queue number (daily reset concept - by max in table)
    cursor.execute("SELECT MAX(no_antrian) FROM antrian")
    result = cursor.fetchone()
    maxno = result[0] if result[0] else 0
    next_no = maxno + 1

    now = datetime.now()
    cursor.execute(
        """INSERT INTO antrian (nama, telepon, email, no_antrian, loket, status, waktu_daftar)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (nama, telepon, email, next_no, None, "Menunggu", now.strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()

    # Save to Excel
    append_to_excel(nama, telepon, email, next_no, now.strftime("%Y-%m-%d %H:%M:%S"))

    return render_template("tiket.html", nama=nama, email=email, telepon=telepon, no_antrian=next_no, waktu=now.strftime("%d/%m/%Y %H:%M"))


# =========================
# API ENDPOINTS
# =========================
@app.route("/api/data-antrian")
def api_data_antrian():
    conn = get_db()
    cursor = conn.cursor()

    result = {"loket": {}, "menunggu": [], "last_called": None}

    # Data per loket (active call)
    for i in range(1, 26):
        cursor.execute(
            "SELECT no_antrian, nama FROM antrian WHERE loket=? AND status='Dipanggil' ORDER BY id DESC LIMIT 1",
            (i,)
        )
        cur = cursor.fetchone()
        if cur:
            result["loket"][f"loket{i}"] = {"no": cur["no_antrian"], "nama": cur["nama"]}
        else:
            result["loket"][f"loket{i}"] = {"no": "-", "nama": ""}

    # Waiting queue (limited to 50 for display performance)
    cursor.execute("SELECT no_antrian, nama FROM antrian WHERE status='Menunggu' ORDER BY id LIMIT 50")
    rows = cursor.fetchall()
    result["menunggu"] = [{"no": row["no_antrian"], "nama": row["nama"]} for row in rows]

    # Last called (for notification animation)
    cursor.execute(
        "SELECT no_antrian, loket, nama FROM antrian WHERE status='Dipanggil' ORDER BY waktu_dipanggil DESC LIMIT 1"
    )
    last = cursor.fetchone()
    if last:
        result["last_called"] = {
            "no": last["no_antrian"],
            "loket": last["loket"],
            "nama": last["nama"]
        }

    # Stats
    cursor.execute("SELECT COUNT(*) FROM antrian WHERE status='Menunggu'")
    result["total_menunggu"] = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM antrian WHERE status='Selesai'")
    result["total_selesai"] = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM antrian")
    result["total_antrian"] = cursor.fetchone()[0]

    conn.close()
    return jsonify(result)


@app.route("/api/interviewer-data")
def api_interviewer_data():
    if session.get("role") != "interviewer":
        return jsonify({"error": "unauthorized"}), 401

    loket = session["loket"]
    conn = get_db()
    cursor = conn.cursor()

    # Current being called
    cursor.execute(
        "SELECT * FROM antrian WHERE loket=? AND status='Dipanggil' ORDER BY id DESC LIMIT 1",
        (loket,)
    )
    called = cursor.fetchone()

    # Waiting count
    cursor.execute("SELECT COUNT(*) FROM antrian WHERE status='Menunggu'")
    menunggu = cursor.fetchone()[0]

    # Completed by this loket today
    cursor.execute("SELECT COUNT(*) FROM antrian WHERE loket=? AND status='Selesai'", (loket,))
    selesai = cursor.fetchone()[0]

    # History of this loket
    cursor.execute(
        "SELECT * FROM antrian WHERE loket=? ORDER BY id DESC LIMIT 10",
        (loket,)
    )
    history = [dict(row) for row in cursor.fetchall()]

    conn.close()

    data = {
        "current": dict(called) if called else None,
        "menunggu": menunggu,
        "selesai": selesai,
        "history": history
    }
    return jsonify(data)


# =========================
# AUTH
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        user = request.form.get("username", "").strip()
        pw = request.form.get("password", "")

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username=?", (user,))
        data = cursor.fetchone()

        if data and check_password_hash(data["password_hash"], pw):
            # Check if this loket is already logged in by another session
            if data["role"] == "interviewer" and data["loket"]:
                cursor.execute("SELECT * FROM active_sessions WHERE loket=?", (data["loket"],))
                active = cursor.fetchone()
                if active:
                    conn.close()
                    error = f"Loket {data['loket']} sudah digunakan oleh user lain. Silakan logout terlebih dahulu."
                    return render_template("login.html", error=error)

                # Register this session as active
                import uuid
                session_id = str(uuid.uuid4())
                cursor.execute(
                    "INSERT INTO active_sessions (loket, session_id, login_time) VALUES (?, ?, ?)",
                    (data["loket"], session_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                )
                conn.commit()
                session["session_id"] = session_id

            conn.close()
            session["role"] = data["role"]
            session["loket"] = data["loket"]
            session["user"] = user
            session["nama_display"] = data["nama_display"]
            if data["role"] == "admin":
                return redirect("/admin")
            else:
                return redirect("/interviewer")
        else:
            conn.close()
        error = "Username atau password salah!"

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    # Remove active session for this loket
    loket = session.get("loket")
    session_id = session.get("session_id")
    if loket and session_id:
        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM active_sessions WHERE loket=? AND session_id=?", (loket, session_id))
            conn.commit()
            conn.close()
        except Exception:
            pass
    session.clear()
    return redirect("/login")


# =========================
# INTERVIEWER
# =========================
@app.route("/interviewer")
def interviewer():
    if session.get("role") != "interviewer":
        return redirect("/login")
    loket = session["loket"]
    nama_display = session.get("nama_display", f"Loket {loket}")
    return render_template("pewawancara.html", loket=loket, nama_display=nama_display)


@app.route("/panggil-berikutnya", methods=["POST"])
def panggil_berikutnya():
    if session.get("role") != "interviewer":
        return jsonify({"error": "unauthorized"}), 401

    loket = session["loket"]
    nama_pewawancara = session.get("nama_display", f"Loket {loket}")
    conn = get_db()
    cursor = conn.cursor()

    # Mark any currently called as 'Dilewati' if not completed
    cursor.execute(
        "UPDATE antrian SET status='Dilewati' WHERE loket=? AND status='Dipanggil'",
        (loket,)
    )

    # Get next waiting
    cursor.execute("SELECT * FROM antrian WHERE status='Menunggu' ORDER BY id LIMIT 1")
    row = cursor.fetchone()

    if row:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            "UPDATE antrian SET status='Dipanggil', loket=?, waktu_dipanggil=?, nama_pewawancara=? WHERE id=?",
            (loket, now, nama_pewawancara, row["id"])
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True, "no_antrian": row["no_antrian"], "nama": row["nama"]})
    
    conn.commit()
    conn.close()
    return jsonify({"success": False, "message": "Tidak ada antrian menunggu"})


@app.route("/selesai-wawancara", methods=["POST"])
def selesai_wawancara():
    if session.get("role") != "interviewer":
        return jsonify({"error": "unauthorized"}), 401

    loket = session["loket"]
    conn = get_db()
    cursor = conn.cursor()

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        "UPDATE antrian SET status='Selesai', waktu_selesai=? WHERE loket=? AND status='Dipanggil'",
        (now, loket)
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True})


@app.route("/panggil-ulang", methods=["POST"])
def panggil_ulang():
    if session.get("role") != "interviewer":
        return jsonify({"error": "unauthorized"}), 401

    loket = session["loket"]
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM antrian WHERE loket=? AND status='Dipanggil' ORDER BY id DESC LIMIT 1",
        (loket,)
    )
    row = cursor.fetchone()
    conn.close()

    if row:
        return jsonify({"success": True, "no_antrian": row["no_antrian"], "nama": row["nama"]})
    return jsonify({"success": False, "message": "Tidak ada antrian yang dipanggil"})


# =========================
# ADMIN
# =========================
@app.route("/admin")
def admin():
    if session.get("role") != "admin":
        return redirect("/login")
    return render_template("admin.html")


@app.route("/api/admin-data")
def api_admin_data():
    if session.get("role") != "admin":
        return jsonify({"error": "unauthorized"}), 401

    conn = get_db()
    cursor = conn.cursor()

    # Pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    per_page = min(per_page, 200)  # Max 200 per page
    search = request.args.get('search', '', type=str).strip()
    offset = (page - 1) * per_page

    # Queue data with pagination and optional search
    if search:
        search_param = f"%{search}%"
        cursor.execute(
            "SELECT COUNT(*) FROM antrian WHERE nama LIKE ? OR CAST(no_antrian AS TEXT) LIKE ? OR telepon LIKE ? OR email LIKE ?",
            (search_param, search_param, search_param, search_param)
        )
        total_filtered = cursor.fetchone()[0]
        cursor.execute(
            "SELECT * FROM antrian WHERE nama LIKE ? OR CAST(no_antrian AS TEXT) LIKE ? OR telepon LIKE ? OR email LIKE ? ORDER BY id DESC LIMIT ? OFFSET ?",
            (search_param, search_param, search_param, search_param, per_page, offset)
        )
    else:
        cursor.execute("SELECT COUNT(*) FROM antrian")
        total_filtered = cursor.fetchone()[0]
        cursor.execute("SELECT * FROM antrian ORDER BY id DESC LIMIT ? OFFSET ?", (per_page, offset))

    antrian = [dict(row) for row in cursor.fetchall()]
    total_pages = max(1, (total_filtered + per_page - 1) // per_page)

    # Stats (using single optimized query)
    cursor.execute("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status='Menunggu' THEN 1 ELSE 0 END) as menunggu,
            SUM(CASE WHEN status='Dipanggil' THEN 1 ELSE 0 END) as dipanggil,
            SUM(CASE WHEN status='Selesai' THEN 1 ELSE 0 END) as selesai,
            SUM(CASE WHEN status='Dilewati' THEN 1 ELSE 0 END) as dilewati
        FROM antrian
    """)
    stats_row = cursor.fetchone()
    total = stats_row["total"] or 0
    menunggu = stats_row["menunggu"] or 0
    dipanggil = stats_row["dipanggil"] or 0
    selesai = stats_row["selesai"] or 0
    dilewati = stats_row["dilewati"] or 0

    # Per loket stats (optimized: 2 queries instead of 75)
    loket_stats = {}
    # Initialize all lokets
    for i in range(1, 26):
        loket_stats[f"loket{i}"] = {"selesai": 0, "dipanggil": 0, "current": None}

    # Get selesai counts per loket in one query
    cursor.execute("SELECT loket, COUNT(*) as cnt FROM antrian WHERE status='Selesai' AND loket IS NOT NULL GROUP BY loket")
    for row in cursor.fetchall():
        key = f"loket{row['loket']}"
        if key in loket_stats:
            loket_stats[key]["selesai"] = row["cnt"]

    # Get current dipanggil per loket in one query
    cursor.execute("""
        SELECT loket, no_antrian, nama FROM antrian 
        WHERE status='Dipanggil' AND loket IS NOT NULL 
        ORDER BY id DESC
    """)
    seen_lokets = set()
    for row in cursor.fetchall():
        key = f"loket{row['loket']}"
        if key in loket_stats and key not in seen_lokets:
            loket_stats[key]["dipanggil"] = 1
            loket_stats[key]["current"] = {"no": row["no_antrian"], "nama": row["nama"]}
            seen_lokets.add(key)

    conn.close()

    return jsonify({
        "antrian": antrian,
        "stats": {
            "menunggu": menunggu,
            "dipanggil": dipanggil,
            "selesai": selesai,
            "dilewati": dilewati,
            "total": total
        },
        "loket_stats": loket_stats,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total_filtered,
            "total_pages": total_pages
        }
    })


@app.route("/admin/panggil/<int:loket>", methods=["POST"])
def admin_panggil(loket):
    if session.get("role") != "admin":
        return jsonify({"error": "unauthorized"}), 401

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM antrian WHERE status='Menunggu' ORDER BY id LIMIT 1")
    row = cursor.fetchone()

    if row:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            "UPDATE antrian SET status='Dipanggil', loket=?, waktu_dipanggil=? WHERE id=?",
            (loket, now, row["id"])
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True, "no_antrian": row["no_antrian"]})

    conn.close()
    return jsonify({"success": False, "message": "Tidak ada antrian menunggu"})


@app.route("/admin/update/<int:id>/<status>", methods=["POST"])
def admin_update(id, status):
    if session.get("role") != "admin":
        return jsonify({"error": "unauthorized"}), 401

    conn = get_db()
    cursor = conn.cursor()

    if status == "Selesai":
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("UPDATE antrian SET status=?, waktu_selesai=? WHERE id=?", (status, now, id))
    else:
        cursor.execute("UPDATE antrian SET status=? WHERE id=?", (status, id))

    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/admin/reset", methods=["POST"])
def admin_reset():
    if session.get("role") != "admin":
        return jsonify({"error": "unauthorized"}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM antrian")
    cursor.execute("DELETE FROM active_sessions")
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/admin/export-excel")
def admin_export_excel():
    if session.get("role") != "admin":
        return redirect("/login")

    if openpyxl is None:
        return "openpyxl tidak tersedia", 500

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM antrian ORDER BY no_antrian")
    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Antrian Wawancara"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["No Antrian", "Nama", "Telepon", "Email", "Status", "Loket", "Waktu Daftar", "Waktu Dipanggil", "Waktu Selesai", "Pewawancara"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        data = [
            row["no_antrian"], row["nama"], row["telepon"] or "-", row["email"] or "-",
            row["status"], row["loket"] or "-", row["waktu_daftar"] or "-",
            row["waktu_dipanggil"] or "-", row["waktu_selesai"] or "-",
            row["nama_pewawancara"] or "-"
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    export_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "export_antrian.xlsx")
    wb.save(export_path)

    from flask import send_file
    return send_file(export_path, as_attachment=True, download_name=f"antrian_wawancara_{get_today_str()}.xlsx")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
